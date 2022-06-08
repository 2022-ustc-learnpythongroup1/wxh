from openpyxl import Workbook

book_lists =[[['b','90','56','e','f'] ,], ]
# book1 = ['b','90','56','e','f'] 
# book_lists.append(book1)
book_tag_lists = ['个人管理']
wb=Workbook()
ws=[]
for i in range(len(book_tag_lists)):
    ws.append(wb.create_sheet(title=book_tag_lists[i])) #utf8->unicode
for i in range(len(book_tag_lists)): 
    ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
    count=1
    for bl in book_lists[i]:
        ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
        count+=1
save_path='testbook_list'
for i in range(len(book_tag_lists)):
    save_path+=('-'+book_tag_lists[i])
save_path+='.xlsx'
wb.save(save_path)